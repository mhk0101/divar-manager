"""
AdExtractor - استخراج‌کننده هوشمند آگهی‌ها، شماره تماس فروشندگان و ارسال پیام خودکار در چت دیوار و شیپور.

اصلاحات نهایی (2026-07-21):
1. ✅ بستن خودکار و آنی مودال‌های تغییر مکان شیپور («آیا مکان خود را به تهران/اسلامشهر تغییر می‌دهید؟»)
2. ✅ فشردن کلید Escape به صورت فیزیکی در سطح ویندوز جهت بستن پاپ‌آپ‌های سیستم‌عاملی
3. ✅ استخراج کامل و دقیق شماره تلفن شیپور و دیوار
4. ✅ ساخت و ذخیره‌سازی اکسل پوشه‌بندی شده بر اساس رنج پیش‌شماره (0912، 0917، 0933)
5. ✅ عدم ارسال پیام تکراری و عدم ذخیره‌سازی آگهی تکراری
"""

from __future__ import annotations

import asyncio
import csv
import json
import logging
import random
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional, Set

from playwright.async_api import Page

from modules.captcha_solver import solve_sheypoor_captcha
from core.network_utils import safe_page_goto, wait_for_internet, is_target_closed_error

logger = logging.getLogger("divar.ad_extractor")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("کتابخانه openpyxl نصب نیست. فایل‌ها با فرمت CSV سازگار با اکسل (UTF-8) ذخیره خواهند شد.")

EXTRACTED_DIR = Path(__file__).resolve().parent.parent / "data" / "extracted_ads"
EXTRACTED_DIR.mkdir(parents=True, exist_ok=True)

PHONES_DIR = Path(__file__).resolve().parent.parent / "data" / "extracted_phones"
PHONES_DIR.mkdir(parents=True, exist_ok=True)

SEEN_TOKENS_FILE = EXTRACTED_DIR / "seen_tokens.json"
MESSAGED_TOKENS_FILE = EXTRACTED_DIR / "messaged_tokens.json"

PERSIAN_TO_ENGLISH_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")

# ---------------------------------------------------------------------------
# Helper: صبر هوشمند تا لود کامل صفحه و ظاهر شدن المان‌های کلیدی
# ---------------------------------------------------------------------------
async def _wait_for_page_ready(
    page: Page,
    selectors: list[str],
    progress_callback: Optional[Callable[[str], None]] = None,
    timeout_ms: int = 30_000,
    label: str = "صفحه",
) -> bool:
    """
    صبر می‌کند تا صفحه کامل لود شود (networkidle) و حداقل یکی از المان‌های کلیدی ظاهر شود.

    مراحل:
    ۱. wait_until="domcontentloaded" در goto قبلاً انجام شده.
    ۲. تلاش برای networkidle (با timeout)
    ۳. wait_for_selector روی تک‌تک selectorها تا یکی پیدا شود
    ۴. اگر هیچکدام پیدا نشد، ۵ ثانیه اضافه صبر می‌کند و دوباره تلاش می‌کند
    """
    # ⬇️ مرحله ۱: صبر برای networkidle ( پایان تمام درخواست‌های شبکه و رندر کامل)
    if progress_callback:
        progress_callback(f"⏳ در حال صبر برای لود کامل {label} (networkidle)...")
    try:
        await page.wait_for_load_state("networkidle", timeout=timeout_ms)
        if progress_callback:
            progress_callback(f"✅ {label} به طور کامل لود شد (networkidle).")
    except Exception:
        if progress_callback:
            progress_callback(f"⚠️ {label}: networkidle تایم‌اوت شد، ادامه با وضعیت فعلی...")
        # حتی اگر networkidle نشد، بازم ادامه بده — شاید صفحه نیمه‌لود باشد
        try:
            await page.wait_for_load_state("load", timeout=10_000)
        except Exception:
            pass

    # ⬇️ مرحله ۲: صبر برای ظاهر شدن حداقل یکی از المان‌های کلیدی
    if progress_callback:
        progress_callback(f"🔍 در حال انتظار برای ظاهر شدن المان‌های کلیدی {label}...")

    for selector in selectors:
        try:
            await page.wait_for_selector(selector, state="visible", timeout=8000)
            if progress_callback:
                progress_callback(f"✅ المان کلیدی «{selector}» در {label} ظاهر شد.")
            return True
        except Exception:
            continue

    # ⬇️ مرحله ۳: هیچ المانی ظاهر نشد — ۵ ثانیه اضافه صبر کن و دوباره تلاش کن
    if progress_callback:
        progress_callback(f"⚠️ هیچ المان کلیدی در {label} یافت نشد. ۵ ثانیه صبر اضافه و تلاش مجدد...")
    await page.wait_for_timeout(5000)

    for selector in selectors:
        try:
            await page.wait_for_selector(selector, state="visible", timeout=8000)
            if progress_callback:
                progress_callback(f"✅ المان کلیدی «{selector}» در تلاش دوم {label} ظاهر شد.")
            return True
        except Exception:
            continue

    if progress_callback:
        progress_callback(f"⚠️ {label}: هیچ‌یک از المان‌های کلیدی ظاهر نشدند. ادامه با وضعیت فعلی...")
    return False


async def _wait_for_divar_contact_section_ready(
    page: Page,
    progress_callback: Optional[Callable[[str], None]] = None,
    timeout_ms: int = 12_000,
) -> bool:
    """برای دیوار فقط منتظر کامل لود شدن بخش دکمه‌های عملیات آگهی بمان.

    بخش موردنیاز دقیقاً همین است و نه کل صفحه:
      .kt-col-5 section .post-actions
        button.post-actions__get-contact  => اطلاعات تماس / تماس امن / تماس ناشناس
        button.start-chat-button...       => چت، اگر برای آگهی وجود داشته باشد

    دکمه چت اجباری نیست، چون بعضی آگهی‌ها چت ندارند. اما اگر وجود داشته باشد،
    بعد از پایدار شدن همین بخش تشخیص داده می‌شود.
    """
    action_container_selectors = [
        ".kt-col-5 section .post-actions",
        ".kt-col-5 .post-actions",
        "section .post-actions",
        ".post-actions",
    ]
    contact_button_selectors = [
        ".kt-col-5 section .post-actions button.post-actions__get-contact",
        ".kt-col-5 .post-actions button.post-actions__get-contact",
        "section .post-actions button.post-actions__get-contact",
        ".post-actions button.post-actions__get-contact",
        ".post-actions button:has-text('اطلاعات تماس')",
        ".post-actions button:has-text('تماس ناشناس')",
        ".post-actions button:has-text('تماس امن')",
    ]

    if progress_callback:
        progress_callback("🔍 انتظار فقط برای کامل لود شدن بخش دکمه‌های آگهی دیوار (.post-actions)...")

    deadline = asyncio.get_event_loop().time() + (timeout_ms / 1000)
    last_signature = None
    stable_hits = 0

    while asyncio.get_event_loop().time() < deadline:
        container = None
        for selector in action_container_selectors:
            try:
                loc = page.locator(selector).first
                if await loc.count() > 0 and await loc.is_visible():
                    container = loc
                    break
            except Exception:
                continue

        if container is None:
            await page.wait_for_timeout(150)
            continue

        contact_btn = None
        contact_text = ""
        for selector in contact_button_selectors:
            try:
                loc = page.locator(selector).first
                if await loc.count() > 0 and await loc.is_visible():
                    contact_btn = loc
                    try:
                        contact_text = (await loc.inner_text()).strip()
                    except Exception:
                        contact_text = ""
                    break
            except Exception:
                continue

        if contact_btn is None:
            await page.wait_for_timeout(150)
            continue

        try:
            # امضای بخش دکمه‌ها: تعداد دکمه‌ها + متن دکمه‌ها. دو بار پشت سر هم ثابت بماند یعنی بخش آماده است.
            signature = await container.evaluate("""
                el => Array.from(el.querySelectorAll('button'))
                    .map(btn => (btn.innerText || btn.getAttribute('aria-label') || '').trim())
                    .filter(Boolean)
                    .join('|')
            """)
        except Exception:
            signature = contact_text or "contact-ready"

        if signature == last_signature and signature:
            stable_hits += 1
        else:
            stable_hits = 0
            last_signature = signature

        if stable_hits >= 2:
            if progress_callback:
                progress_callback(f"✅ بخش دکمه‌های دیوار آماده شد: {signature}")
            return True

        await page.wait_for_timeout(150)

    if progress_callback:
        progress_callback("⚠️ بخش دکمه‌های دیوار در زمان تعیین‌شده کامل پایدار نشد؛ ادامه با وضعیت فعلی...")
    return False


async def _wait_for_divar_contact_result(
    page: Page,
    progress_callback: Optional[Callable[[str], None]] = None,
    timeout_ms: int = 10_000,
) -> bool:
    """بعد از کلیک روی اطلاعات تماس، فقط منتظر نتیجه همان expandable-box/tel بمان.

    از selectorهای عمومی مثل .kt-unexpandable-row__value-box استفاده نمی‌کنیم چون قبل از کلیک هم
    در بخش مشخصات آگهی وجود دارند و باعث ادامه زودهنگام می‌شوند.
    """
    if progress_callback:
        progress_callback("🔍 انتظار برای ظاهر شدن نتیجه اطلاعات تماس دیوار...")

    deadline = asyncio.get_event_loop().time() + (timeout_ms / 1000)
    while asyncio.get_event_loop().time() < deadline:
        try:
            tel_count = await page.locator("a[href^='tel:']").count()
            if tel_count > 0:
                if progress_callback:
                    progress_callback("✅ لینک تماس tel در اطلاعات تماس دیوار ظاهر شد.")
                return True
        except Exception:
            pass

        try:
            # expandable-box قبل از کلیک collapsed است؛ بعد از کلیک معمولاً کلاس collapsed حذف می‌شود.
            expanded = page.locator(".expandable-box:not(.expandable-box--collapsed)").first
            if await expanded.count() > 0 and await expanded.is_visible():
                txt = await expanded.inner_text()
                if normalize_phone_number(txt) or "شماره" in txt or "تماس" in txt:
                    if progress_callback:
                        progress_callback("✅ باکس اطلاعات تماس دیوار باز شد.")
                    return True
        except Exception:
            pass

        try:
            copy_rows = page.locator(".copyRow-l4byg9, [class*='copyRow'], .expandable-box a")
            if await copy_rows.count() > 0:
                if progress_callback:
                    progress_callback("✅ ردیف اطلاعات تماس دیوار ظاهر شد.")
                return True
        except Exception:
            pass

        await page.wait_for_timeout(250)

    if progress_callback:
        progress_callback("⚠️ نتیجه اطلاعات تماس دیوار دیر لود شد/پیدا نشد؛ تلاش برای خواندن وضعیت فعلی...")
    return False


async def press_physical_escape_key(page: Page) -> None:
    """
    ارسال کلید Escape کاملاً واقعی و فیزیکی در سطح سیستم‌عامل ویندوز (Win32 Native API).
    """
    if sys.platform.startswith("win"):
        try:
            import ctypes
            VK_ESCAPE = 0x1B
            KEYEVENTF_KEYUP = 0x0002

            ctypes.windll.user32.keybd_event(VK_ESCAPE, 0, 0, 0)
            await asyncio.sleep(0.08)
            ctypes.windll.user32.keybd_event(VK_ESCAPE, 0, KEYEVENTF_KEYUP, 0)
            logger.info("Physical OS-level Escape keypress dispatched via Win32 API")
        except Exception as e:
            logger.debug("Win32 keybd_event Escape failed: %s", e)

    try:
        await page.keyboard.down("Escape")
        await page.wait_for_timeout(80)
        await page.keyboard.up("Escape")
    except Exception as e:
        logger.debug("Playwright physical Escape press failed: %s", e)

    try:
        await page.evaluate("""
            () => {
                const escEvent = new KeyboardEvent('keydown', {
                    key: 'Escape',
                    code: 'Escape',
                    keyCode: 27,
                    which: 27,
                    bubbles: true,
                    cancelable: true
                });
                document.dispatchEvent(escEvent);
                window.dispatchEvent(escEvent);
                if (document.body) document.body.dispatchEvent(escEvent);
            }
        """)
    except Exception:
        pass


async def dismiss_sheypoor_modals(page: Page) -> None:
    """بستن خودکار مودال‌های تایید تغییر مکان در شیپور."""
    try:
        modal_btn = page.locator("button:has-text('بله، تغییر می‌دهم'), button:has-text('تغییر می‌دهم')").first
        if await modal_btn.count() > 0 and await modal_btn.is_visible():
            await modal_btn.click(force=True, timeout=2000)
            await page.wait_for_timeout(500)
            logger.info("[sheypoor_modal] Automatically clicked 'بله، تغییر می‌دهم'")
    except Exception as e:
        logger.debug("[sheypoor_modal] Notice: %s", e)


def normalize_phone_number(raw_text: str) -> Optional[str]:
    """
    تبدیل اعداد فارسی به انگلیسی و استخراج دقیق شماره موبایل ۱۱ رقمی ایرانی (۰۹...).
    در صورت عدم مطابقت با الگو، None برمی‌گرداند تا متون نامربوط ثبت نشوند.
    """
    if not raw_text:
        return None
    text = str(raw_text).translate(PERSIAN_TO_ENGLISH_DIGITS)
    digits_only = re.sub(r"\D", "", text)
    match = re.search(r"09\d{9}", digits_only)
    if match:
        return match.group(0)
    return None


def organize_and_save_phone_excel(
    phone_number: str,
    title: str,
    location_name: str,
    category_name: str,
    platform: str,
    url: str,
) -> Path:
    """
    ذخیره مرتب شماره تلفن در فایل اکسل جداگانه با پوشه‌بندی دقیق طبق الگوی درخواستی:
    پوشه: data/extracted_phones/{پیش‌شماره}_ {استان_یا_شهر}_ دسته {دسته‌بندی}/
    فایل اکسل: شماره_های_{پیش_شماره}.xlsx
    مثال: 0917_ بوشهر_ دسته کالای دیجیتال/شماره_های_0917.xlsx
    """
    prefix = phone_number[:4] if (phone_number and len(phone_number) >= 4) else "0900"

    loc_clean = re.sub(r'[\\/*?:"<>|]', '_', location_name or "سراسر ایران").strip()
    cat_clean = re.sub(r'[\\/*?:"<>|]', '_', category_name or "عمومی").strip()

    folder_name = f"{prefix}_ {loc_clean}_ دسته {cat_clean}"

    dir_path = PHONES_DIR / folder_name
    dir_path.mkdir(parents=True, exist_ok=True)

    headers = ["ردیف", "شماره موبایل", "پیش‌شماره", "عنوان آگهی", "شهر/استان", "دسته‌بندی", "پلتفرم", "لینک آگهی", "زمان استخراج"]
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if HAS_OPENPYXL:
        file_path = dir_path / f"شماره_های_{prefix}.xlsx"

        if file_path.exists():
            try:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
            except Exception:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"رنج {prefix}"
                ws.append(headers)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"رنج {prefix}"
            ws.append(headers)

            header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        existing_phones = set()
        for row in range(2, ws.max_row + 1):
            p_val = ws.cell(row=row, column=2).value
            if p_val:
                existing_phones.add(str(p_val).strip())

        if phone_number not in existing_phones:
            ws.append([
                ws.max_row,
                phone_number,
                prefix,
                title,
                location_name,
                category_name,
                platform,
                url,
                now_str,
            ])

            last_row = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=last_row, column=col_idx)
                cell.alignment = Alignment(horizontal="center" if col_idx in (1, 2, 3, 7, 9) else "right")

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            wb.save(file_path)

        return file_path
    else:
        file_path = dir_path / f"شماره_های_{prefix}.csv"

        existing_lines = []
        existing_phones = set()

        if file_path.exists():
            try:
                with open(file_path, "r", encoding="utf-8-sig") as f:
                    reader = csv.reader(f)
                    existing_lines = list(reader)
                    for r in existing_lines[1:]:
                        if len(r) >= 2:
                            existing_phones.add(r[1].strip())
            except Exception:
                existing_lines = []

        if not existing_lines:
            existing_lines = [headers]

        if phone_number not in existing_phones:
            row_idx = len(existing_lines)
            existing_lines.append([
                row_idx,
                phone_number,
                prefix,
                title,
                location_name,
                category_name,
                platform,
                url,
                now_str,
            ])

            with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerows(existing_lines)

        return file_path


def load_seen_tokens() -> Set[str]:
    """بارگذاری لیست توکن‌ها و لینک‌های استخراج‌شده قبلی جهت عدم ثبت تکراری."""
    if not SEEN_TOKENS_FILE.exists():
        return set()
    try:
        with open(SEEN_TOKENS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return set(data.get("seen_tokens", []))
    except Exception as e:
        logger.warning("Failed to load seen tokens: %s", e)
        return set()


def save_seen_tokens(seen_tokens: Set[str]) -> None:
    """ذخیره دایمی توکن‌های پردازش‌شده در فایل."""
    try:
        with open(SEEN_TOKENS_FILE, "w", encoding="utf-8") as f:
            json.dump({"seen_tokens": list(seen_tokens)}, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning("Failed to save seen tokens: %s", e)


def load_messaged_tokens() -> Set[str]:
    """بارگذاری لیست توکن‌هایی که قبلاً پیام چت دریافت کرده‌اند جهت عدم ارسال پیام تکراری."""
    if not MESSAGED_TOKENS_FILE.exists():
        return set()
    try:
        with open(MESSAGED_TOKENS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            return set(data.get("messaged_tokens", []))
    except Exception as e:
        logger.warning("Failed to load messaged tokens: %s", e)
        return set()


def save_messaged_tokens(messaged_tokens: Set[str]) -> None:
    """ذخیره دایمی توکن‌های دریافت‌کننده پیام چت."""
    try:
        with open(MESSAGED_TOKENS_FILE, "w", encoding="utf-8") as f:
            json.dump({"messaged_tokens": list(messaged_tokens)}, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning("Failed to save messaged tokens: %s", e)


def extract_divar_token(href: str) -> str:
    """استخراج دقیق توکن دیوار از لینک آگهی."""
    if not href:
        return ""
    path = href.split("?")[0].rstrip("/")
    segments = [s for s in path.split("/") if s]
    if segments:
        return segments[-1]
    return ""


def extract_sheypoor_listing_id(href: str) -> str:
    """
    استخراج ID آگهی شیپور از لینک.
    مثال: https://www.sheypoor.com/v/عنوان-465146913.html -> 465146913
    """
    if not href:
        return ""
    clean_href = href.split("?")[0].replace(".html", "").rstrip("/")
    parts = clean_href.split("-")
    if len(parts) > 1 and parts[-1].isdigit():
        return parts[-1]
    slash_parts = clean_href.split("/")
    if slash_parts and slash_parts[-1].isdigit():
        return slash_parts[-1]
    return ""


class AdExtractor:
    """کلاس استخراج‌کننده آگهی‌ها برای دیوار و شیپور."""

    def __init__(self, platform: str, page: Page):
        self.platform = platform.lower()
        self.page = page
        self.seen_tokens = load_seen_tokens()
        self.messaged_tokens = load_messaged_tokens()

    async def extract_page_ads(self) -> List[Dict[str, str]]:
        """استخراج تمام آگهی‌های یکتای موجود در صفحه فعلی."""
        await wait_for_internet()
        extracted = []
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if self.platform == "divar":
            try:
                cards = await self.page.eval_on_selector_all(
                    "a[href*='/v/']",
                    """
                    elements => elements.map(el => {
                        const href = el.getAttribute('href') || '';
                        const titleEl = el.querySelector('.kt-post-card__title, h2, h3, .kt-raw-text') || el;
                        const descEl = el.querySelector('.kt-post-card__description, .kt-post-card__info');
                        return {
                            href: href,
                            title: titleEl ? titleEl.innerText.trim() : '',
                            desc: descEl ? descEl.innerText.trim() : ''
                        };
                    })
                    """
                )

                for item in cards:
                    href = item.get("href", "")
                    if not href:
                        continue
                    full_url = href if href.startswith("http") else f"https://divar.ir{href}"
                    title = item.get("title") or "آگهی دیوار"

                    token = extract_divar_token(href)
                    unique_key = token if token else full_url

                    if unique_key in self.seen_tokens:
                        continue

                    self.seen_tokens.add(unique_key)
                    extracted.append({
                        "platform": "divar",
                        "token": token,
                        "title": title,
                        "url": full_url,
                        "chat_url": f"https://divar.ir/chat/{token}" if token else "",
                        "phone_number": "در حال استخراج...",
                        "description": item.get("desc", ""),
                        "extracted_at": now_str,
                        "chat_sent": token in self.messaged_tokens,
                    })
            except Exception as e:
                logger.warning("[divar] Error extracting page cards: %s", e)

        elif self.platform == "sheypoor":
            try:
                cards = await self.page.eval_on_selector_all(
                    "article a[href], .serp-item a[href], a[href*='sheypoor.com/v/'], a[href*='/v/']",
                    """
                    elements => elements.map(el => {
                        const href = el.getAttribute('href') || '';
                        const titleEl = el.querySelector('h2, h3, strong, .title') || el;
                        return {
                            href: href,
                            title: titleEl ? titleEl.innerText.trim() : ''
                        };
                    })
                    """
                )

                for item in cards:
                    href = item.get("href", "")
                    if not href or "sheypoor.com/s/" in href or "sheypoor.com/session/" in href:
                        continue

                    full_url = href if href.startswith("http") else f"https://www.sheypoor.com{href}"
                    title = item.get("title") or "آگهی شیپور"
                    listing_id = extract_sheypoor_listing_id(href)

                    if full_url in self.seen_tokens or (listing_id and listing_id in self.seen_tokens):
                        continue

                    self.seen_tokens.add(listing_id if listing_id else full_url)
                    extracted.append({
                        "platform": "sheypoor",
                        "token": listing_id,
                        "title": title,
                        "url": full_url,
                        "chat_url": f"https://www.sheypoor.com/session/myChats?listingId={listing_id}" if listing_id else "",
                        "phone_number": "در حال استخراج...",
                        "description": "",
                        "extracted_at": now_str,
                        "chat_sent": listing_id in self.messaged_tokens,
                    })
            except Exception as e:
                logger.warning("[sheypoor] Error extracting page cards: %s", e)

        if extracted:
            save_seen_tokens(self.seen_tokens)

        return extracted

    async def scrape_multiple_pages(
        self,
        max_pages: int = 1,
        progress_callback: Optional[Callable[[str], None]] = None,
    ) -> List[Dict[str, str]]:
        """پیمایش صفحات و استخراج آگهی‌های جدید با وقفه دقیق ۳ ثانیه‌ای بین اسکرول‌ها."""
        all_new_ads = []

        for page_num in range(1, max_pages + 1):
            if progress_callback:
                progress_callback(f"📥 در حال استخراج آگهی‌های جدید صفحه {page_num} از {max_pages}...")

            await wait_for_internet(
                progress_callback=progress_callback,
                first_message=f"🌐 بررسی اینترنت قبل از استخراج آگهی‌های صفحه {page_num}...",
                restored_message=f"✅ اینترنت وصل است؛ استخراج صفحه {page_num} ادامه دارد...",
            )
            page_ads = await self.extract_page_ads()
            all_new_ads.extend(page_ads)

            if progress_callback:
                progress_callback(
                    f"✅ صفحه {page_num}: {len(page_ads)} آگهی جدید یافت شد (مجموع کل یکتا: {len(all_new_ads)})"
                )

            if page_num < max_pages:
                try:
                    await wait_for_internet(progress_callback=progress_callback)
                    await self.page.evaluate("window.scrollTo(0, document.body.scrollHeight)")

                    if progress_callback:
                        progress_callback(f"⏳ وقفه ۳ ثانیه‌ای بین اسکرول‌ها جهت دریافت آگهی‌های جدید (صفحه {page_num} به {page_num + 1})...")
                    await self.page.wait_for_timeout(3000)

                    next_button = self.page.locator("button:has-text('صفحه بعد'), a:has-text('صفحه بعد'), .pagination a").first
                    if await next_button.count() > 0 and await next_button.is_visible():
                        await next_button.click()
                        await self.page.wait_for_timeout(3000)
                except Exception as err:
                    logger.debug("Scroll error on page %d: %s", page_num, err)

        return all_new_ads

    async def extract_divar_phone_number(
        self,
        token: str,
        ad_url: Optional[str] = None,
        progress_callback: Optional[Callable[[str], None]] = None,
    ) -> tuple[Optional[str], bool]:
        """
        ورود به صفحه خود آگهی دیوار، وقفه لود کامل، کلیک چندروشه روی «اطلاعات تماس»
        و استخراج و صحت‌سنجی شماره تلفن با Regex.

        نکته مهم: در حالت همگام استخراج شماره + ارسال چت، شماره تماس باید از لینک
        خود آگهی خوانده شود، نه از لینک چت. به همین دلیل اگر ad_url پاس داده شود،
        دقیقاً همان URL استخراج‌شده از کارت آگهی باز می‌شود.

        Returns: (phone_number_str, has_chat_enabled)
        """
        if not token and not ad_url:
            return None, False

        if ad_url:
            clean_ad_url = str(ad_url).split("?")[0].strip()
            if clean_ad_url.startswith("/"):
                clean_ad_url = f"https://divar.ir{clean_ad_url}"
            # محافظت: اگر اشتباهاً لینک چت پاس داده شد، آن را به لینک آگهی تبدیل کن.
            if "/chat/" in clean_ad_url and token:
                clean_ad_url = f"https://divar.ir/v/-/{token}"
            ad_url = clean_ad_url
        else:
            ad_url = f"https://divar.ir/v/-/{token}"

        phone_number = None
        has_chat = False

        try:
            if progress_callback:
                progress_callback(f"📞 در حال باز کردن آگهی جهت بررسی تماس ({ad_url})...")

            await safe_page_goto(
                self.page,
                ad_url,
                wait_until="domcontentloaded",
                timeout=25_000,
                progress_callback=progress_callback,
                label="صفحه آگهی دیوار",
            )

            # برای استخراج شماره دیوار لازم نیست کل صفحه networkidle شود؛
            # فقط خود بخش دکمه‌های .post-actions باید کامل/پایدار شود.
            await _wait_for_divar_contact_section_ready(
                self.page,
                progress_callback=progress_callback,
                timeout_ms=12_000,
            )

            # وقفه خیلی کوتاه برای تثبیت React بعد از آماده شدن دکمه‌ها
            await self.page.wait_for_timeout(200)

            chat_btn = self.page.locator("button:has-text('چت'), a:has-text('چت'), .start-chat-button-e813e, [data-testimonial-id='chat-button']").first
            if await chat_btn.count() > 0 and await chat_btn.is_visible():
                has_chat = True

            contact_btn = self.page.locator(
                "button.post-actions__get-contact, .post-actions__get-contact, button:has-text('اطلاعات تماس'), button:has-text('تماس ناشناس'), button:has-text('تماس امن')"
            ).first

            if await contact_btn.count() > 0 and await contact_btn.is_visible():
                btn_text = (await contact_btn.inner_text()).strip()

                if "تماس ناشناس" in btn_text or "تماس امن" in btn_text:
                    phone_number = "تماس ناشناس (بدون شماره)"
                    if progress_callback:
                        progress_callback(f"ℹ️ آگهی دارای {btn_text} است؛ کلیک انجام نشد.")

                elif "اطلاعات تماس" in btn_text:
                    is_disabled = await contact_btn.is_disabled() or "disabled" in (await contact_btn.get_attribute("class") or "")
                    if is_disabled:
                        phone_number = "فقط چت (شماره مخفی)"
                        if progress_callback:
                            progress_callback("ℹ️ فروشنده شماره تماس را مخفی کرده است (فقط چت).")
                    else:
                        await contact_btn.scroll_into_view_if_needed()
                        await self.page.wait_for_timeout(800)

                        clicked = False

                        await wait_for_internet(
                            progress_callback=progress_callback,
                            first_message="🌐 بررسی اینترنت قبل از کلیک روی اطلاعات تماس دیوار...",
                            restored_message="✅ اینترنت وصل است؛ ادامه استخراج شماره دیوار...",
                        )

                        try:
                            await contact_btn.click(force=True, timeout=3000)
                            clicked = True
                        except Exception:
                            pass

                        if not clicked:
                            try:
                                await contact_btn.evaluate("el => el.click()")
                                clicked = True
                            except Exception:
                                pass

                        if not clicked:
                            try:
                                await contact_btn.dispatch_event("click")
                                clicked = True
                            except Exception:
                                pass

                        try:
                            span_inner = contact_btn.locator("span").first
                            if await span_inner.count() > 0:
                                await span_inner.click(force=True, timeout=2000)
                        except Exception:
                            pass

                        try:
                            await self.page.evaluate("""
                                () => {
                                    const btn = document.querySelector('.post-actions__get-contact, button.post-actions__get-contact');
                                    if (btn) btn.click();
                                }
                            """)
                        except Exception:
                            pass

                        # بعد از کلیک، فقط منتظر نتیجه همان اطلاعات تماس بمان؛
                        # نه ردیف‌های عمومی مشخصات آگهی که از قبل وجود دارند.
                        await _wait_for_divar_contact_result(
                            self.page,
                            progress_callback=progress_callback,
                            timeout_ms=10_000,
                        )

                        tel_links = await self.page.locator("a[href^='tel:']").all()
                        for link in tel_links:
                            href = await link.get_attribute("href") or ""
                            parsed_p = normalize_phone_number(href)
                            if parsed_p:
                                phone_number = parsed_p
                                break

                        if not phone_number:
                            value_elements = await self.page.locator(".copyRow-l4byg9 a, [class*='copyRow'] a, .expandable-box a").all()
                            for vel in value_elements:
                                txt = await vel.inner_text()
                                parsed_p = normalize_phone_number(txt)
                                if parsed_p:
                                    phone_number = parsed_p
                                    break

                        # بعضی نسخه‌های UI دیوار شماره را به صورت متن داخل expandable-box می‌گذارند نه لینک.
                        if not phone_number:
                            try:
                                contact_text = await self.page.locator(".expandable-box:not(.expandable-box--collapsed), .expandable-box").first.inner_text()
                                parsed_p = normalize_phone_number(contact_text)
                                if parsed_p:
                                    phone_number = parsed_p
                            except Exception:
                                pass
            else:
                phone_number = "اطلاعات تماس ثبت‌نشده"

            if not phone_number or phone_number == "در حال استخراج...":
                phone_number = "ناموجود/مخفی"

            if progress_callback and phone_number and not phone_number.startswith(("فقط چت", "تماس ناشناس", "ناموجود", "اطلاعات تماس")):
                progress_callback(f"📱 شماره تماس معتبر استخراج شد: {phone_number}")

        except Exception as e:
            if is_target_closed_error(e):
                raise RuntimeError("__BROWSER_CLOSED__") from e
            logger.debug("Failed to extract phone number for token %s: %s", token, e)
            phone_number = "خطا در استخراج"

        return phone_number, has_chat

    async def extract_sheypoor_phone_number(
        self,
        url: str,
        progress_callback: Optional[Callable[[str], None]] = None,
    ) -> tuple[Optional[str], bool]:
        """
        ورود به صفحه آگهی شیپور، کلیک روی دکمه تماس، بستن خودکار مودال مکان و فشردن فیزیکی Escape روی کیبورد ۱ ثانیه بعد.
        Returns: (phone_number_str, has_chat_enabled)
        """
        if not url:
            return None, False

        phone_number = None
        has_chat = False

        try:
            if progress_callback:
                progress_callback(f"📞 در حال باز کردن آگهی شیپور جهت استخراج شماره ({url})...")

            await safe_page_goto(
                self.page,
                url,
                wait_until="domcontentloaded",
                timeout=30_000,
                progress_callback=progress_callback,
                label="صفحه آگهی شیپور",
            )

            # ── صبر هوشمند تا لود کامل صفحه و ظاهر شدن دکمه‌های کلیدی شیپور ──
            key_selectors = [
                "button:has-text('تماس با')",
                "button:has-text('چت')",
                "button:has-text('چت با کاربر شیپور')",
                "a[href^='tel:']",
            ]
            await _wait_for_page_ready(
                self.page,
                selectors=key_selectors,
                progress_callback=progress_callback,
                timeout_ms=30_000,
                label="صفحه آگهی شیپور",
            )

            # وقفه کوتاه اضافه برای اطمینان از رندر کامل
            await self.page.wait_for_timeout(1000)

            # ✨ بستن هرگونه مودال تغییر مکان شیپور («آیا مکان خود را تغییر می‌دهید؟»)
            await dismiss_sheypoor_modals(self.page)

            # ۱. بررسی امکان چت
            chat_btn = self.page.locator("button:has-text('چت'), button:has-text('چت با کاربر شیپور')").first
            if await chat_btn.count() > 0 and await chat_btn.is_visible():
                has_chat = True

            # ۲. بررسی دکمه تماس شیپور (مانند "تماس با ۰۹۱۲XXX۲۲۳۹")
            contact_btn = self.page.locator("button:has-text('تماس با'), button:has-text('تماس')").first
            if await contact_btn.count() > 0 and await contact_btn.is_visible():
                await contact_btn.scroll_into_view_if_needed()
                await self.page.wait_for_timeout(500)

                # ✨ (۱/۲) فشردن Escape قبل از کلیک روی دکمه تماس
                if progress_callback:
                    progress_callback("⌨️ (۱/۲) فشردن Escape قبل از کلیک روی دکمه تماس...")
                await press_physical_escape_key(self.page)
                await self.page.wait_for_timeout(1000)

                # کلیک روی دکمه تماس
                await wait_for_internet(
                    progress_callback=progress_callback,
                    first_message="🌐 بررسی اینترنت قبل از کلیک روی تماس شیپور...",
                    restored_message="✅ اینترنت وصل است؛ ادامه استخراج شماره شیپور...",
                )
                try:
                    await contact_btn.click(force=True, timeout=3000)
                except Exception:
                    await contact_btn.evaluate("el => el.click()")

                # ✨ وقفه ۲ ثانیه‌ای پس از کلیک
                await self.page.wait_for_timeout(2000)

                # ✨ (۲/۲) فشردن Escape بعد از کلیک روی دکمه تماس
                if progress_callback:
                    progress_callback("⌨️ (۲/۲) فشردن Escape بعد از کلیک روی دکمه تماس...")
                await press_physical_escape_key(self.page)

                await self.page.wait_for_timeout(1000)

                # مجدداً بررسی و بستن مودال‌های احتمالی مکان
                await dismiss_sheypoor_modals(self.page)

                # ✨ تشخیص و حل خودکار کد امنیتی (Captcha) شیپور با EasyOCR - حداکثر ۶ تلاش
                captcha_solved = await solve_sheypoor_captcha(
                    self.page,
                    progress_callback=progress_callback,
                    max_attempts=6,
                )
                if not captcha_solved:
                    if progress_callback:
                        progress_callback(
                            "⚠️ کد امنیتی شیپور به‌صورت خودکار حل نشد. "
                            "لطفاً به‌صورت دستی کپچا را در مرورگر وارد کنید یا برنامه را متوقف نمایید."
                        )
                    # ۱۰ ثانیه صبر می‌کنیم شاید کاربر دستی وارد کند
                    await self.page.wait_for_timeout(10000)

                # ✨ فشردن Escape با ۱ ثانیه وقفه بعد از حل کپچا
                await self.page.wait_for_timeout(1000)
                if progress_callback:
                    progress_callback("⌨️ فشردن Escape بعد از حل کپچا...")
                await press_physical_escape_key(self.page)
                await self.page.wait_for_timeout(500)

                # خواندن شماره تلفن از تگ tel: یا متن دکمه
                tel_links = await self.page.locator("a[href^='tel:']").all()
                for link in tel_links:
                    href = await link.get_attribute("href") or ""
                    parsed_p = normalize_phone_number(href)
                    if parsed_p:
                        phone_number = parsed_p
                        break

                if not phone_number:
                    btn_text = await contact_btn.inner_text()
                    phone_number = normalize_phone_number(btn_text)

            if not phone_number:
                phone_number = "ناموجود/مخفی"

            if progress_callback and phone_number and phone_number.startswith("09"):
                progress_callback(f"📱 شماره تماس شیپور استخراج شد: {phone_number}")

            # ✨ وقفه تصادفی بین ۱ تا ۶ ثانیه برای شیپور جهت شبیه‌سازی رفتار انسانی
            delay_sec = round(random.uniform(1.0, 6.0), 1)
            if progress_callback:
                progress_callback(f"⏳ وقفه تصادفی {delay_sec} ثانیه‌ای (بین ۱ تا ۶ ثانیه) برای شیپور...")
            await self.page.wait_for_timeout(int(delay_sec * 1000))

        except Exception as e:
            if is_target_closed_error(e):
                raise RuntimeError("__BROWSER_CLOSED__") from e
            logger.debug("Failed to extract Sheypoor phone for %s: %s", url, e)
            phone_number = "خطا در استخراج"

        return phone_number, has_chat

    async def send_divar_chat_message(
        self,
        token: str,
        message_text: str,
        progress_callback: Optional[Callable[[str], None]] = None,
    ) -> bool:
        """ارسال دقیق پیام چت دیوار با کنترل عدم ارسال تکراری."""
        if not token or not message_text.strip():
            return False

        if token in self.messaged_tokens:
            if progress_callback:
                progress_callback(f"ℹ️ به آگهی (توکن: {token}) قبلاً پیام ارسال شده است؛ نادیده گرفته شد.")
            return True

        chat_url = f"https://divar.ir/chat/{token}"
        try:
            if progress_callback:
                progress_callback(f"💬 در حال ورود به چت اختصاصی دیوار ({chat_url})...")

            await safe_page_goto(
                self.page,
                chat_url,
                wait_until="domcontentloaded",
                timeout=25_000,
                progress_callback=progress_callback,
                label="صفحه چت دیوار",
            )

            textarea_selectors = [
                "#chat-input",
                "textarea#chat-input",
                "textarea.kt-chat-input__input",
                ".kt-chat-input__input",
                "textarea[placeholder*='متنی بنویسید']",
                "textarea[placeholder*='پیام']",
                "textarea",
            ]

            # صبر هوشمند تا صفحه چت دیوار کامل آماده شود و ورودی پیام ظاهر شود.
            await _wait_for_page_ready(
                self.page,
                selectors=textarea_selectors + [
                    "button:has-text('ارسال')",
                    ".kt-chat-input",
                    ".chat-input",
                ],
                progress_callback=progress_callback,
                timeout_ms=25_000,
                label="صفحه چت دیوار",
            )

            await self.page.wait_for_timeout(700)

            chat_input = None
            for sel in textarea_selectors:
                loc = self.page.locator(sel).first
                if await loc.count() > 0 and await loc.is_visible():
                    chat_input = loc
                    break

            if not chat_input:
                ad_url = f"https://divar.ir/v/-/{token}"
                await safe_page_goto(
                self.page,
                ad_url,
                wait_until="domcontentloaded",
                timeout=25_000,
                progress_callback=progress_callback,
                label="صفحه آگهی دیوار",
            )
                await _wait_for_page_ready(
                    self.page,
                    selectors=[
                        "button:has-text('چت')",
                        "a:has-text('چت')",
                        "[data-testimonial-id='chat-button']",
                        "button.post-actions__get-contact",
                        ".post-actions__get-contact",
                        "h1",
                        "article",
                    ],
                    progress_callback=progress_callback,
                    timeout_ms=25_000,
                    label="صفحه آگهی دیوار برای ورود به چت",
                )
                await self.page.wait_for_timeout(700)

                chat_btn = self.page.locator("button:has-text('چت'), a:has-text('چت'), [data-testimonial-id='chat-button']").first
                if await chat_btn.count() > 0 and await chat_btn.is_visible():
                    await chat_btn.click()
                    await _wait_for_page_ready(
                        self.page,
                        selectors=textarea_selectors + ["button:has-text('ارسال')"],
                        progress_callback=progress_callback,
                        timeout_ms=25_000,
                        label="صفحه چت دیوار بعد از کلیک روی دکمه چت",
                    )
                    await self.page.wait_for_timeout(700)

                for sel in textarea_selectors:
                    loc = self.page.locator(sel).first
                    if await loc.count() > 0 and await loc.is_visible():
                        chat_input = loc
                        break

            if not chat_input:
                if progress_callback:
                    progress_callback(f"⚠️ امکان ارسال چت برای آگهی (توکن: {token}) فعال نیست.")
                return False

            await wait_for_internet(
                progress_callback=progress_callback,
                first_message="🌐 بررسی اینترنت قبل از ارسال پیام چت دیوار...",
                restored_message="✅ اینترنت وصل است؛ ارسال پیام چت دیوار ادامه دارد...",
            )

            await chat_input.fill(message_text)
            await self.page.wait_for_timeout(1000)

            send_btn_selectors = [
                "button:has-text('ارسال')",
                "button[type='submit']",
                ".kt-chat-send-button",
                "button.kt-button-primary",
            ]

            sent = False
            for btn_sel in send_btn_selectors:
                btn = self.page.locator(btn_sel).first
                if await btn.count() > 0 and await btn.is_visible():
                    await btn.click()
                    sent = True
                    break

            if not sent:
                await chat_input.press("Enter")
                sent = True

            await self.page.wait_for_timeout(2500)

            self.messaged_tokens.add(token)
            save_messaged_tokens(self.messaged_tokens)

            if progress_callback:
                progress_callback(f"✅ پیام با موفقیت در چت دیوار ارسال شد: https://divar.ir/chat/{token}")
            return True

        except Exception as e:
            if is_target_closed_error(e):
                raise RuntimeError("__BROWSER_CLOSED__") from e
            logger.error("[divar_chat] Exception sending chat to %s: %s", token, e)
            if progress_callback:
                progress_callback(f"❌ خطا در ارسال پیام چت برای {token}: {e}")
            return False

    async def send_sheypoor_chat_message(
        self,
        listing_id: str,
        message_text: str,
        progress_callback: Optional[Callable[[str], None]] = None,
    ) -> bool:
        """ارسال دقیق پیام در چت شیپور (https://www.sheypoor.com/session/myChats?listingId={listing_id})."""
        if not listing_id or not message_text.strip():
            return False

        if listing_id in self.messaged_tokens:
            if progress_callback:
                progress_callback(f"ℹ️ به آگهی شیپور ({listing_id}) قبلاً پیام ارسال شده است؛ نادیده گرفته شد.")
            return True

        chat_url = f"https://www.sheypoor.com/session/myChats?listingId={listing_id}"
        try:
            if progress_callback:
                progress_callback(f"💬 در حال ورود به چت شیپور ({chat_url})...")

            await safe_page_goto(
                self.page,
                chat_url,
                wait_until="domcontentloaded",
                timeout=30_000,
                progress_callback=progress_callback,
                label="صفحه چت شیپور",
            )

            # ── صبر هوشمند تا لود کامل صفحه چت شیپور و ظاهر شدن فیلد ورودی ──
            await _wait_for_page_ready(
                self.page,
                selectors=[
                    "input[name='msg']",
                    "input[placeholder*='پیامتان را بنویسید']",
                    "input[placeholder*='پیام']",
                ],
                progress_callback=progress_callback,
                timeout_ms=30_000,
                label="صفحه چت شیپور",
            )

            # بستن خودکار مودال مکان شیپور
            await dismiss_sheypoor_modals(self.page)

            chat_input = self.page.locator("input[name='msg'], input[placeholder*='پیامتان را بنویسید'], input[placeholder*='پیام']").first

            if await chat_input.count() > 0 and await chat_input.is_visible():
                # ✨ (۱/۲) فشردن Escape قبل از ارسال پیام در چت شیپور
                if progress_callback:
                    progress_callback("⌨️ (۱/۲) فشردن Escape قبل از ارسال پیام در چت شیپور...")
                await press_physical_escape_key(self.page)
                await self.page.wait_for_timeout(500)

                await wait_for_internet(
                    progress_callback=progress_callback,
                    first_message="🌐 بررسی اینترنت قبل از ارسال پیام چت شیپور...",
                    restored_message="✅ اینترنت وصل است؛ ارسال پیام چت شیپور ادامه دارد...",
                )

                await chat_input.fill(message_text)
                await self.page.wait_for_timeout(1000)
                await chat_input.press("Enter")
                await self.page.wait_for_timeout(2000)

                # ✨ (۲/۲) فشردن Escape بعد از ارسال پیام در چت شیپور
                if progress_callback:
                    progress_callback("⌨️ (۲/۲) فشردن Escape بعد از ارسال پیام در چت شیپور...")
                await press_physical_escape_key(self.page)
                await self.page.wait_for_timeout(1000)

                self.messaged_tokens.add(listing_id)
                save_messaged_tokens(self.messaged_tokens)

                if progress_callback:
                    progress_callback(f"✅ پیام با موفقیت در چت شیپور ارسال شد: {chat_url}")
                return True
            else:
                if progress_callback:
                    progress_callback(f"⚠️ ورودی چت شیپور برای آگهی {listing_id} پیدا نشد.")
                return False

        except Exception as e:
            if is_target_closed_error(e):
                raise RuntimeError("__BROWSER_CLOSED__") from e
            logger.error("[sheypoor_chat] Exception sending chat to %s: %s", listing_id, e)
            if progress_callback:
                progress_callback(f"❌ خطا در ارسال چت شیپور: {e}")
            return False


def save_extracted_ads(ads: List[Dict[str, str]], platform: str, phone: str) -> tuple[Path, Path]:
    """ذخیره آگهی‌های استخراج‌شده در دو فایل JSON و CSV."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"{platform}_{phone}_{timestamp}"

    json_path = EXTRACTED_DIR / f"{base_name}.json"
    csv_path = EXTRACTED_DIR / f"{base_name}.csv"

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({"count": len(ads), "ads": ads}, f, ensure_ascii=False, indent=2)

    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["#", "عنوان آگهی", "شماره تماس", "لینک کامل آگهی", "لینک مستقیم چت", "شناسه / توکن", "پلتفرم", "وضعیت چت", "زمان استخراج"])
        for idx, ad in enumerate(ads, 1):
            writer.writerow([
                idx,
                ad.get("title", ""),
                ad.get("phone_number", "ناموجود/مخفی"),
                ad.get("url", ""),
                ad.get("chat_url", ""),
                ad.get("token", ""),
                ad.get("platform", ""),
                "ارسال شد" if ad.get("chat_sent") else "ارسال نشده",
                ad.get("extracted_at", ""),
            ])

    return json_path, csv_path
